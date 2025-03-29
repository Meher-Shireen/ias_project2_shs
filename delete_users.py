import sqlite3

conn = sqlite3.connect("authentication.db")
cursor = conn.cursor()

# Delete all records from users table
cursor.execute("DELETE FROM users")
conn.commit()
conn.close()

print("All users deleted from database.")


#excel
from openpyxl import load_workbook

wb = load_workbook("users.xlsx")
ws = wb.active

# Keep headers, delete all data rows
ws.delete_rows(2, ws.max_row)
wb.save("users.xlsx")

print("All user records removed from Excel.")
