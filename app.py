from flask import Flask, request, jsonify
import os
import sqlite3
import openpyxl  # To handle Excel
from openpyxl import load_workbook, Workbook
from authentication import hash_password, encrypt_aes, decrypt_aes, verify_password
from authentication import generate_biometric
from database import init_db
from crypto_utils import generate_rsa_keypair, encrypt_private_key
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
from key_manager import get_aes_key
import base64
import rsa
from cryptography.hazmat.primitives import padding
from datetime import datetime
app = Flask(__name__)

# Initialize database (run only once)
init_db()

# Secret key for AES encryption (should be stored securely, like in an env file)
AES_KEY = get_aes_key()


EXCEL_FILE = "users.xlsx"  # Ensure this file exists before running
DEVICE_EXCEL_FILE = "devices.xlsx"  # Ensure this file exists before running

def save_to_excel(user_id, username, hashed_password, encrypted_biometric, user_public_key, encrypted_user_private_key):
    """Saves user details to an Excel file."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)  # Load existing file
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()  # Create new workbook if not exists
        sheet = wb.active
        sheet.append(["ID", "Username", "Hashed Password", "Encrypted Biometric", "user_public_key", "encrypted_user_private_key"])  # Header row

    # Append new user data
    sheet.append([user_id, username, hashed_password, encrypted_biometric, user_public_key, encrypted_user_private_key])
    wb.save(EXCEL_FILE)  # Save changes

def save_device_to_excel(device_id, device_name, user_id, encrypted_secret, device_public_key, encrypted_device_private_key):
    """Saves device details to an Excel file."""
    try:
        wb = openpyxl.load_workbook(DEVICE_EXCEL_FILE)  # Load existing file
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()  # Create new workbook if not exists
        sheet = wb.active
        sheet.append(["Device ID", "Device Name", "User ID", "Encrypted Secret", "device_public_key", "encrypted_device_private_key"])  # Header row

    # Append new device data
    sheet.append([device_id, device_name, user_id, encrypted_secret, device_public_key, encrypted_device_private_key])
    wb.save(DEVICE_EXCEL_FILE)  # Save changes

def register_user_to_db(username, password, biometric, user_public_key, encrypted_user_private_key):
    """Helper function to register a user in the database."""
    hashed_password, salt = hash_password(password)
    encrypted_biometric = encrypt_aes(biometric, AES_KEY)
    
    try:
        conn = sqlite3.connect("authentication.db")
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO users (username, password, biometric, user_public_key, encrypted_user_private_key) VALUES (?, ?, ?, ?, ?)",
            (username, hashed_password.hex() + ":" + salt.hex(), encrypted_biometric, user_public_key, encrypted_user_private_key)
        )
        conn.commit()
        
        user_id = cursor.lastrowid  # Get inserted user's ID
        conn.close()

        # Save user details to Excel
        save_to_excel(user_id, username, hashed_password.hex() + ":" + salt.hex(), encrypted_biometric, user_public_key, encrypted_user_private_key)

        return user_id
    except sqlite3.IntegrityError:
        return False

def register_device_to_db(user_id, device_name, secret, device_public_key, encrypted_device_private_key):
    """Registers a device in the database."""
    encrypted_secret = encrypt_aes(secret, AES_KEY)
    
    try:
        conn = sqlite3.connect("authentication.db")
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO devices (user_id, device_name, secret, device_public_key, encrypted_device_private_key) VALUES (?, ?, ?, ?, ?)",
            (user_id, device_name, encrypted_secret, device_public_key, encrypted_device_private_key)
        )
        conn.commit()
        
        device_id = cursor.lastrowid  # Get inserted device's ID
        conn.close()

        # Save device details to Excel
        save_device_to_excel(device_id, device_name, user_id, encrypted_secret, device_public_key, encrypted_device_private_key)

        return True
    except sqlite3.IntegrityError:
        return False

@app.route('/register', methods=['POST'])
def register_user():
    """Registers a single user via API request."""
    data = request.json  # Expecting JSON input
    username = data.get('username')
    password = data.get('password')
    biometric = generate_biometric()

    private_key, user_public_key = generate_rsa_keypair()
    encrypted_user_private_key = encrypt_private_key(private_key)
    
    if not username or not password or not biometric:
        return jsonify({'error': 'Missing required fields'}), 400
    
    user_id = register_user_to_db(username, password, biometric, user_public_key, encrypted_user_private_key)
    if user_id:
        return jsonify({'message': 'User registered successfully', 'user_id': user_id}), 201
    else:
        return jsonify({'error': 'Username already exists'}), 409

import uuid

@app.route('/register_device', methods=['POST'])
def register_device():
    """Registers a new device for a user, generating secret automatically."""
    data = request.json
    device_name = data.get('device_name')
    user_id = data.get('user_id')

    private_key, device_public_key = generate_rsa_keypair()
    encrypted_device_private_key = encrypt_private_key(private_key)

    if not user_id or not device_name:
        return jsonify({'error': 'Missing required fields'}), 400
    
    # Automatically generate secret instead of requiring it in the request
    secret = f"{device_name}_{user_id}_{uuid.uuid4()}"

    if register_device_to_db(user_id, device_name, secret, device_public_key, encrypted_device_private_key):
        return jsonify({'message': 'Device registered successfully'}), 201
    else:
        return jsonify({'error': 'Device registration failed'}), 409

@app.route('/send_message', methods=['POST'])
def send_message():
    """Handles message sending from device to user via gateway."""
    data = request.json
    device_id = data.get('device_id')
    user_id = data.get('user_id')
    encrypted_message = data.get('encrypted_message')

    if not device_id or not user_id or not encrypted_message:
        return jsonify({'error': 'Missing required fields'}), 400

    # Check if device is registered
    conn = sqlite3.connect("authentication.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM devices WHERE id = ? AND user_id = ?", (device_id, user_id))
    device = cursor.fetchone()
    conn.close()

    if not device:
        return jsonify({'error': 'Device not registered or not linked to user'}), 404

    # Forward the message to the user (in this case, just save it to the database)
    try:
        conn = sqlite3.connect("authentication.db")
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO messages (device_id, user_id, encrypted_message) VALUES (?, ?, ?)",
            (device_id, user_id, encrypted_message)
        )
        conn.commit()
        conn.close()
        return jsonify({'message': 'Message sent successfully'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Failed to send message'}), 500



@app.route('/receive_message', methods=['GET'])
def receive_message():
    """Handles message retrieval and decryption by the user."""
    user_id = request.args.get('user_id')

    if not user_id:
        return jsonify({'error': 'Missing user_id'}), 400

    # Retrieve the latest message for the user
    conn = sqlite3.connect("authentication.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM messages WHERE user_id = ? ORDER BY id DESC LIMIT 1", (user_id,))
    message = cursor.fetchone()
    conn.close()

    if not message:
        return jsonify({'error': 'No messages found'}), 404

    encrypted_message = message[3]  # Verify this index is correct for your DB schema

    try:
        decrypted_message = decrypt_aes(encrypted_message, get_aes_key())
        return jsonify({
            'device_id': message[1],  # Verify correct column index
            'decrypted_message': decrypted_message
        }), 200
    except Exception as e:
        print(f"Decryption Error: {e}")  # Debugging output
        return jsonify({'error': 'Decryption failed'}), 500
    
@app.route('/send_message_to_device', methods=['POST'])
def send_message_to_device():
    print("\n=== Received request ===")  # Debug
    print("Headers:", request.headers)   # Debug
    try:
        data = request.get_json()
        print("Raw data received:", data)  # Debug
        
        if not data:
            print("Error: No JSON data received")  # Debug
            return jsonify({"error": "No JSON data received"}), 400
            
        device_id = data.get("device_id")
        message = data.get("message")
        user_id = data.get("user_id")
        
        print(f"Parsed data - Device: {device_id}, User: {user_id}, Message: {message}")  # Debug

        if not all([device_id, message, user_id]):
            print("Error: Missing fields")  # Debug
            return jsonify({"error": "Missing device_id, message, or user_id"}), 400

        # Encrypt the message
        encrypted_message = encrypt_aes(message, AES_KEY)
        
        # Save message to device_messages table
        conn = sqlite3.connect("authentication.db")
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO device_messages (device_id, user_id, encrypted_message, timestamp) VALUES (?, ?, ?, ?)",
            (device_id, user_id, encrypted_message, datetime.now())
        )
        conn.commit()
        conn.close()

        print("Message processed and saved successfully")  # Debug
        return jsonify({"success": "Message sent to device!"}), 200
        
    except Exception as e:
        print(f"Error in send_message_to_device: {str(e)}")  # Debug
        return jsonify({"error": str(e)}), 500

@app.route('/get_messages_for_device/<int:device_id>', methods=['GET'])
def get_messages_for_device(device_id):
    """Device retrieves its messages (decrypted with its private key)."""
    secret = request.args.get("secret")  # Basic auth (replace with better auth in production)

    conn = sqlite3.connect("authentication.db")
    cursor = conn.cursor()

    # 1. Verify device and fetch its private key
    cursor.execute("""
        SELECT encrypted_device_private_key 
        FROM devices 
        WHERE id = ? AND secret = ?
    """, (device_id, secret))
    device = cursor.fetchone()
    if not device:
        conn.close()
        return jsonify({"error": "Invalid device or secret"}), 403

    # 2. Decrypt the device's private key using AES
    decrypted_private_key_pem = decrypt_aes(device[0], AES_KEY)
    device_private_key = rsa.PrivateKey.load_pkcs1(decrypted_private_key_pem.encode())

    # 3. Fetch messages
    cursor.execute("""
        SELECT encrypted_message, timestamp 
        FROM device_messages 
        WHERE device_id = ? 
        ORDER BY timestamp DESC
    """, (device_id,))
    encrypted_messages = cursor.fetchall()
    conn.close()

    # 4. Decrypt messages
    decrypted_messages = []
    for enc_msg, timestamp in encrypted_messages:
        try:
            decrypted = rsa.decrypt(base64.b64decode(enc_msg), device_private_key).decode()
            decrypted_messages.append({"message": decrypted, "timestamp": timestamp})
        except Exception as e:
            decrypted_messages.append({"error": str(e), "timestamp": timestamp})

    return jsonify({"messages": decrypted_messages})

if __name__ == '__main__':
    app.run(debug=True)